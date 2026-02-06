from odoo import models, fields, api
from odoo.exceptions import UserError
import io
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


class RentPaymentReportWizard(models.TransientModel):
    _name = 'rent.payment.report.wizard'
    _description = 'Rent Payment Report Wizard'

    date_from = fields.Date(string='From Date')
    date_to = fields.Date(string='To Date')
    status = fields.Selection([
        ('all', 'All'),
        ('unpaid', 'Unpaid'),
        ('paid', 'Paid')
    ], string='Status', default='all')

    def generate_excel_report(self):
        # Fetch records based on filters
        domain = []
        if self.date_from:
            domain.append(('payment_date', '>=', self.date_from))
        if self.date_to:
            domain.append(('payment_date', '<=', self.date_to))
        if self.status != 'all':
            domain.append(('status', '=', self.status))

        payments = self.env['rent.payment'].search(domain)

        if not payments:
            raise UserError("No records found for the selected criteria.")

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Rent Payment Report"
        #
        # Define styles
        header_font = Font(bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(horizontal='center', vertical='center')
        #
        # # Write headers
        headers = ['Payment Reference', 'Lease', 'Payment Date', 'Amount Paid', 'Status', 'Note']
        for col, header in enumerate(headers, start=2):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border
        #
        # # Write data
        for row, payment in enumerate(payments, start=2):
            ws.cell(row=row, column=1, value=payment.name).border = border
            ws.cell(row=row, column=2, value=payment.lease_id.name).border = border
            ws.cell(row=row, column=3, value=payment.payment_date.strftime('%Y-%m-%d')).border = border
            ws.cell(row=row, column=4, value=payment.amount_paid).border = border
            ws.cell(row=row, column=5,
                    value=dict(payment._fields['status'].selection).get(payment.status)).border = border
            ws.cell(row=row, column=6, value=payment.note or '').border = border

            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column].width = adjusted_width

            # Save the workbook to a binary stream
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # Encode the file in base64
            excel_data = base64.b64encode(output.read())
            output.close()

            # Create attachment
            attachment = self.env['ir.attachment'].create({
                'name': 'Rent_Payment_Report.xlsx',
                'type': 'binary',
                'datas': excel_data,
                'res_model': self._name,
                'res_id': self.id,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            })

            # Return action to download the file
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?download=true',
                'target': 'self',
            }